import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
import os
from openpyxl.styles import Alignment
from summarizer import generate_darpe_summary
from datetime import datetime

load_dotenv()

def is_not_expired(deadline_str):
    if not deadline_str:
        return True
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(str(deadline_str).strip(), fmt) >= datetime.today()
        except ValueError:
            continue
    return True

# 1. Start a session
session = requests.Session()

login_url = "https://darpe.me/darpe-login.php" # The URL the form submits to
target_url = "https://darpe.me/tenders-and-grants/" # The page you actually want to scrape

user_field_name = os.getenv("USER_FIELD_NAME")
username = os.getenv("USER_NAME")
# pass_field_name = os.getenv("PASSWORD_FIELD_NAME")
password = os.getenv("PASSWORD")


# 2. Create your payload using the exact field names the site expects
login_data = {
    "log": username,
    "pwd": password,
    "wp-submit": "Sign in",
    "redirect_to": "https://darpe.me"
    # Sometimes you need a hidden token here too, like an anti-CSRF token
}

# 3. Send the login request
print("Logging in...")
login_response = session.post(login_url, data=login_data)

detail_response = session.get(target_url)

# 5. Parse the protected HTML
soup = BeautifulSoup(detail_response.text, "html.parser")

BASE_URL = "https://darpe.me"
LISTING_URL = "https://darpe.me/tenders-and-grants/"  # replace with actual listing URL

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

#PART 1: SCRAPES FROM LIST ON TABLE PAGE AND CONVERTS TO CSV FILE

# collapse whitespace and normalize scraped text
def clean_text(text):
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()

#searches for "client name" within title cell. used in extract_listing_rows function
def extract_client_name(td):
    text = clean_text(td.get_text(" ", strip=True))
    match = re.search(r"Client Name\s*:\s*(.+)", text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

#extracts rows
def extract_listing_rows(html):
    soup = BeautifulSoup(html, "html.parser")
    # print(soup)
    rows = soup.select("tr.whiteBackground, tr.graybackground, tr.grayBackground")
    results = []

    for row in rows:
        tds = row.find_all("td")
        if not tds:
            continue
        
        # 0) type (tender or grant)
        row_text = clean_text(row.get_text(" ", strip=True)).lower()
        if "grant" in row_text:
            listing_type = "Grant"
        elif "tender" in row_text:
            listing_type = "Tender"
        else:
            listing_type = "Other"


        # 1) title 
        title_link = row.select_one("a[href*='darpe-entries']")
        title = clean_text(title_link.get_text()) if title_link else ""
        detail_page_url = urljoin(BASE_URL, title_link["href"]) if title_link and title_link.has_attr("href") else ""
        
        # 2) donor name from "Client Name : ..."
        donor_name = ""
        if title_link:
            parent_td = title_link.find_parent("td")
            if parent_td:
                donor_name = extract_client_name(parent_td)

        # 3) deadline (uses index)
        deadline = clean_text(tds[2].get_text(" ", strip=True)) if len(tds) > 2 else ""
        
        if not is_not_expired(deadline):
            continue

        # 4) sector/services (uses index)
        focus_sector = clean_text(tds[3].get_text(" ", strip=True)) if len(tds) > 3 else ""

        # 5) geography (uses index)
        geographic_area = clean_text(tds[4].get_text(" ", strip=True)) if len(tds) > 4 else ""

        # Access inner link
        info_soup = None
        try:
            res = session.get(detail_page_url, headers=HEADERS)  # use session!
            res.raise_for_status()
            info_soup = BeautifulSoup(res.text, "html.parser")
        except Exception as e:
            print(f"Error crawling {detail_page_url}: {e}")
        
        # 6) attachments
        attachment_urls = []
        if info_soup:
            # find the "Attachments" heading
            attachments_heading = info_soup.find(lambda tag: tag.name and "Attachments" in tag.get_text() and tag.name in ["p", "h2", "h3", "h4", "strong", "b"])
            if attachments_heading:
                parent_li = attachments_heading.find_parent("li")
                if parent_li:
                    attachment_urls = [a["href"] for a in parent_li.find_all("a", href=True)]

        # 7) og link
        og_link = ""
        if info_soup:
            bold_p = info_soup.find("p", style=lambda s: s and "font-weight:bold" in s, string=lambda t: t and "Link to original" in t)
            if bold_p:
                parent_li = bold_p.find_parent("li")
                og_anchor = parent_li.find("a", href=True) if parent_li else None
                og_link = og_anchor["href"] if og_anchor else ""

        results.append({
            "type": listing_type,
            "title": title,
            "detail_page_url": detail_page_url,
            "donor_name": donor_name,
            "deadline": deadline,
            "focus_sector": focus_sector,
            "geographic_area": geographic_area,
            "attachments": attachment_urls,
            "original link": og_link,
            "ai_summary": ""
        })

    return results

#makes http get request
def fetch_html(url):
    resp = session.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.text

#FILTER FOR KEYWORDS, TYPES, GEOGRAPHIES

WORKFORCE_KEYWORDS = [
    # Employment and Workforce
    "youth employment",
    "workforce development",
    "employability",
    "job placement",
    "job creation",
    "livelihoods",
    "economic empowerment",
    "economic inclusion",
    "apprenticeship",
    "internship",
    "mentorship",
    "job readiness",
    "job search",
    "labor market activation",
    "economic participation",
    "labor market entry",
    "NEET",
    "work readiness",
    "job seekers",
    "early-career",
    "access to opportunities",
    "reducing inequalities",

    # Skills and Training
    "skills development",
    "vocational training",
    "technical training",
    "soft skills",
    "digital skills",
    "vocational skills",
    "technical skills",
    "green jobs",
    "green skills",
    "TVET",
    "upskilling",
    "reskilling",
    "employability skills",
    "curriculum development",
    "vocational training center",
    "career center",
    "AI skills",
    "climate change",
    "financial literacy",
    "circular economy",
    "higher education",
    "university",
    "educational institutions",
    "life skills",
    "transversal skills",
    "entrepreneurial skills",
    "blended training",

    # Entrepreneurship and Private Sector
    "entrepreneurship",
    "SME development",
    "private sector development",
    "self employment",
    "virtual jobs",
    "income generation",
    "startup incubation",
    "employer engagement",
    "business acceleration",
    "micro entrepreneurship",
    "new business creation",
    "SME",
    "incubation",
    "green entrepreneurship",
    "women entrepreneurship",
    "startup",
    "startup support",
    "financial inclusion",
    "home-based businesses",
    "MSME",
    "microbusiness",
    "freelance",
    "gig work",
    "gig economy",

    # Systems Change
    "capacity building",
    "systems strengthening",
    "framework",
    "action plan",
    "competitiveness",
    "skills gaps",
    "business association",
    "chamber of commerce",
    "industry federation",
]

ALLOWED_TYPES = ["Tender", "Grant"]

ALLOWED_GEOGRAPHIES = [
    "morocco", "algeria", "tunisia", "egypt", "jordan", "palestine",
    "UAE", "united arab emirates", "saudi arabia", "lebanon",
    "bahrain", "syria", "mena"
]

def is_workforce_related(row):
    text = " ".join([
        row.get("type", ""),
        row.get("title", ""),
        row.get("detail_page_url", ""),
        row.get("donor_name", ""),
        row.get("deadline", ""),
        row.get("focus_sector", ""),
        row.get("geographic_area", ""),
    ]).lower()
    return any(keyword.lower() in text for keyword in WORKFORCE_KEYWORDS)

def is_allowed_type(row):
    return row.get("type", "") in ALLOWED_TYPES

def is_in_geography(row):
    text = row.get("geographic_area", "").lower()
    return any(geo in text for geo in ALLOWED_GEOGRAPHIES)

def apply_filters(df):
    df["filter_workforce"] = df.apply(is_workforce_related, axis=1)
    df["filter_type"]       = df.apply(is_allowed_type, axis=1)
    df["filter_geography"]  = df.apply(is_in_geography, axis=1)
    df["passes_all"]        = df["filter_workforce"] & df["filter_type"] & df["filter_geography"]
    return df

def get_total_pages():
    html = fetch_html(LISTING_URL)
    soup = BeautifulSoup(html, "html.parser")
    
    # grab all page number links and find the highest one
    page_links = soup.select("a.page-numbers")
    page_numbers = []
    for link in page_links:
        text = link.get_text(strip=True).replace(",", "")  # removes comma from "1,913"
        if text.isdigit():
            page_numbers.append(int(text))
    
    return max(page_numbers) if page_numbers else 1
total_pages = get_total_pages()

#RUNS THE ACTUAL CODE:

MAX_PAGES = 1 # ← change this to however many pages you want

all_items = []

for page_num in range(1, MAX_PAGES + 1):
    print(f"Scraping page {page_num} of {MAX_PAGES}...")
    
    if page_num == 1:
        page_url = LISTING_URL
    else:
        page_url = f"{BASE_URL}/tenders-and-grants/page/{page_num}/"
    
    html = fetch_html(page_url)
    items = extract_listing_rows(html)
    
    if not items:
        print(f"No items found on page {page_num}, stopping early.")
        break
    
    all_items.extend(items)
    time.sleep(1)

print(f"Total items scraped: {len(all_items)}")

df = pd.DataFrame(all_items)
# df with true or false value whether it matches a requirement
df = apply_filters(df)

# df with only the true values from above
df = df[df["passes_all"] == True]

wrap = Alignment(wrap_text=True, vertical="top")
#CONVERTING TO EXCEL SPREADSHEET
'''wb = Workbook()
ws = wb.active
ws.title = "Tenders & Grants"

# Header row
# headers = ["Title", "Type","Donor Name", "Geographic Area", "Focus Sector", "Deadline", "Source Link", "Original Link", "Attachments",
#            "AI Summary", "Amount (USD)", "Eligibility"]

headers = ["Title", "Post Link", "Donor Name", "Geographic Area", "Focus / Sector", "Grant Link","Deadline", "Description", "AI Summary"]
ws.append(headers)

# Style header row
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF", name="Arial")
    cell.fill = PatternFill("solid", start_color="2E4057")
    cell.alignment = Alignment(horizontal="center")

# Generating AI Summaries
for i, row in df.iterrows():
    text_for_ai = " ".join([
        row["title"],
        row["focus_sector"],
        row["geographic_area"]
    ])

    summary = generate_darpe_summary(text_for_ai)

    df.at[i, "ai_summary"] = summary

    time.sleep(12) # prevents Gemini rate limit

# Data rows
for _, row in df.iterrows():
    ws.append([
        row["title"],
        row["detail_page_url"],
        row["donor_name"],
        row["geographic_area"],
        row["focus_sector"],
        row["original link"],
        row["deadline"],
        ", ".join(row["attachments"]) if row["attachments"] else "",
        row["ai_summary"],
        "",  # Amount USD (cannot find so filler for now)
        ""   # Eligibility (cannot find so filler for now)
    ])

# Auto-fit column widths
# Auto-fit column widths based on content
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    adjusted_width = min(max_length + 4, 80)  # +4 padding, cap at 80
    ws.column_dimensions[col[0].column_letter].width = adjusted_width

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Freeze header row
ws.freeze_panes = "A2"

wb.save("tenders_grants.xlsx")
print("Saved to tenders_grants.xlsx")'''

EXCEL_FILE = os.getenv("EXCEL_FILE")
SHEET_NAME = "darpe"

def _write_headers_and_data(ws, df):
    headers = ["Title", "Type", "Donor Name", "Geographic Area", "Focus Sector",
               "Deadline", "Source Link", "Original Link", "Attachments", "AI Summary"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center")

    for i, row in df.iterrows():
        text_for_ai = " ".join([
            row["title"],
            row["focus_sector"],
            row["geographic_area"]
        ])

        summary = generate_darpe_summary(text_for_ai)

        df.at[i, "ai_summary"] = summary

        time.sleep(12) # prevents Gemini rate limit

    for _, row in df.iterrows():
        ws.append([
            row["title"], row["type"], row["donor_name"], row["geographic_area"],
            row["focus_sector"], row["deadline"], row["detail_page_url"],
            row["original link"],
            ", ".join(row["attachments"]) if row["attachments"] else "",
            row["ai_summary"]
        ])

def _apply_style(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 80)
    ws.freeze_panes = "A2"

def write_styled_sheet(df, excel_file, sheet_name):
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        if sheet_name in wb.sheetnames:
            # Sheet exists — dedup and append only new rows
            existing_df = pd.read_excel(excel_file, sheet_name=sheet_name)
            existing_links = set(existing_df["Source Link"])
            new_rows = df[~df["detail_page_url"].isin(existing_links)]
            if new_rows.empty:
                print("No new grants")
                return
            ws = wb[sheet_name]
            
            for i, row in new_rows.iterrows():
                text_for_ai = " ".join([
                    row["title"],
                    row["focus_sector"],
                    row["geographic_area"]
                ])
                summary = generate_darpe_summary(text_for_ai)
                new_rows.at[i, "ai_summary"] = summary
                time.sleep(12)
                
            for _, row in new_rows.iterrows():
                ws.append([
                    row["title"], row["type"], row["donor_name"], row["geographic_area"],
                    row["focus_sector"], row["deadline"], row["detail_page_url"],
                    row["original link"],
                    ", ".join(row["attachments"]) if row["attachments"] else "",
                    row["ai_summary"]
                ])
            print(f"Added {len(new_rows)} new grants")
        else:
            # File exists, sheet doesn't
            ws = wb.create_sheet(sheet_name)
            _write_headers_and_data(ws, df)
            print(f"Created new sheet '{sheet_name}'")
    else:
        # File doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        _write_headers_and_data(ws, df)
        print("Created new Excel file")

    _apply_style(wb[sheet_name])
    wb.save(excel_file)

write_styled_sheet(df, EXCEL_FILE, SHEET_NAME)
