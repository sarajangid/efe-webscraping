import argparse
import re
import time
import requests
from bs4 import BeautifulSoup
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
from reqs import *
from summarizer import generate_sam_summary
import os
from dotenv import load_dotenv
from datetime import datetime

load_dotenv()
EXCEL_FILE = os.environ["EXCEL_FILE"]

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_COLOR = "1F4E79"
HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

DEFAULT_WIDTHS = {
    "Opportunity ID": 18,
    "Opportunity Type": 22,
    "Title": 40,
    "Donor Name": 28,
    "Geographic Area": 24,
    "Focus / Sector": 28,
    "Application Deadline": 20,
    "Amount Min (USD)": 18,
    "Amount Max (USD)": 18,
    "Eligibility": 42,
    "Eligibility Requirements": 45,
    "Matched Keywords": 32,
    "Source Link": 45,
    "Original Link": 45,
    "Date Posted": 18,
    "Date Scraped": 16,
    "Grant Name": 38,
    "Agency": 28,
    "Due Date": 18,
    "Award Minimum": 18,
    "Award Maximum": 18,
    "Description": 65,
    "Documents": 45,
    "Application Link": 45,
    "AI Summary": 65,
    "Post Link": 50,
    "Funding Amount (USD)": 22,
    "Grant Link": 45,
    "Deadline": 18,
}

def is_not_expired(deadline_str):
    if not deadline_str:
        return True
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(str(deadline_str).strip(), fmt) >= datetime.today()
        except ValueError:
            continue
    return True

def _auto_width(header: str) -> float:
    header = (header or "").strip()
    if not header:
        return 18
    return min(max(len(header) + 4, 18), 45)

def apply_impact_formatting(path: str, sheet_name: str, custom_widths=None):
    wb = openpyxl.load_workbook(path)

    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    widths = {**DEFAULT_WIDTHS, **(custom_widths or {})}

    if ws.max_row >= 1:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            header = "" if cell.value is None else str(cell.value)

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(
                header, _auto_width(header)
            )

        ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = BODY_ALIGNMENT

    ws.freeze_panes = "A2"
    wb.save(path)

BASE_URL = "https://www.developmentaid.org/tenders/search?sort=relevance.desc&searchedText=grants"

def norm(text):
    return re.sub(r"\s+", " ", text or "").strip()

def matches(text, terms):
    """Return terms found in text using whole-word matching.
    Plain substring matching causes false positives:
      "Levant"  matches inside "reLEVANT"  (very common in grant text)
      "MENA"    matches inside "pheMENAl", "MENAce", "aMENAble"
    Word boundaries prevent these."""
    if not text:
        return []
    return [t for t in terms if re.search(r'\b' + re.escape(t) + r'\b', text, re.IGNORECASE)]

def parse_amount(text):
    nums = [a.replace(",", "") for a in re.findall(r"[\$€£]?\s*([\d,]+(?:\.\d+)?)", text or "")]
    return ("", "") if not nums else (nums[0], nums[0]) if len(nums) == 1 else (nums[0], nums[-1])

DETAIL_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
}

def bs_get(soup, *selectors):
    for sel in selectors:
        el = soup.select_one(sel)
        if el:
            return norm(el.get_text(" "))
    return ""

def get(page, sel):
    el = page.query_selector(sel)
    return norm(el.inner_text()) if el else ""

def get_links(page):
    """Return deduplicated list of {href, surface} dicts from a listing page."""
    cards = page.query_selector_all("a[href*='/tenders/']")
    print(f"  Found {len(cards)} cards")
    seen, out = set(), []
    for card in cards:
        try:
            href = card.get_attribute("href") or ""
            if not re.search(r"/tenders/(?:view/|)\d+", href) or href in seen:
                continue
            if not href.startswith("http"): href = "https://www.developmentaid.org" + href
            seen.add(href)
            out.append({"href": href, "surface": norm(card.inner_text())})
        except Exception:
            continue
    return out

def scrape_detail(href, source_url):
    """Fetch a detail page with requests+BS4; return row dict or None if filtered out."""
    time.sleep(1)
    try:
        resp = requests.get(href, headers=DETAIL_HEADERS, timeout=25)
        resp.raise_for_status()
    except Exception as e:
        print(f"    [WARN] Request failed {href}: {e}")
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    # Strip nav/sidebar so geo matching isn't polluted by filter labels
    for el in soup.select("nav, header, footer, aside, [class*='sidebar'], [class*='filter'], [class*='nav']"):
        el.decompose()

    main = soup.select_one("main, [role='main'], article, [class*='tender-detail'], [class*='tender-body'], [class*='detail-content']")
    content = norm(main.get_text(" ")) if main and len(main.get_text()) > 200 else norm(soup.get_text(" "))

    opp_type = bs_get(soup, "[class*='type']", "[class*='Type']", ".badge", ".tag", "[class*='category']")
    if opp_type and "grant" not in opp_type.lower() and "grant" not in content[:2000].lower():
        return None

    geo = bs_get(soup, "[class*='location']", "[class*='Location']", "[class*='country']", "[class*='Country']", "[class*='region']")
    mena_hits = matches(geo, MENA_COUNTRIES) or matches(content[:4000], MENA_COUNTRIES)
    if not mena_hits:
        return None

    kw_hits = matches(content, KEYWORDS)
    if not kw_hits:
        return None

    m = re.search(r"/tenders/(\d+)", href)
    opp_id = m.group(1) if m else re.sub(r"\W", "", href)[-12:]

    title       = bs_get(soup, "h1", "[class*='title'] h1", "[class*='Title']") or "N/A"
    donor       = bs_get(soup, "[class*='donor']", "[class*='Donor']", "[class*='funder']", "[class*='client']", "[class*='organisation']")
    sector      = bs_get(soup, "[class*='sector']", "[class*='Sector']", "[class*='focus']", "[class*='theme']")
    eligibility = bs_get(soup, "[class*='eligib']", "[class*='Eligib']", "[class*='applicant']", "[class*='eligible']")
    deadline    = bs_get(soup, "[class*='deadline']", "[class*='Deadline']", "[class*='closing']", "time[datetime]")
    date_posted = bs_get(soup, "[class*='posted']", "[class*='Published']", "[class*='published']", "time")
    
    if not is_not_expired(deadline):
        return None
    amt_min, amt_max = parse_amount(bs_get(soup, "[class*='amount']", "[class*='Amount']", "[class*='budget']", "[class*='Budget']", "[class*='funding']"))

    return {
        "Opportunity ID":       opp_id,
        "Opportunity Type":     opp_type or "Grant",
        "Title":                title,
        "Donor Name":           donor,
        "Geographic Area":      geo or ", ".join(mena_hits),
        "Focus / Sector":       sector,
        "Application Deadline": deadline,
        "Amount Min (USD)":     amt_min,
        "Amount Max (USD)":     amt_max,
        "Eligibility":          eligibility,
        "Matched Keywords":     "; ".join(kw_hits),
        "Source Link":          source_url,
        "Original Link":        href,
        "Date Posted":          date_posted,
        "AI Summary":           None,
        "_opp_data":            {
            "Title": title, "Donor Name": donor,
            "Geographic Area": geo or ", ".join(mena_hits),
            "Focus / Sector": sector, "Eligibility": eligibility,
            "Amount Max (USD)": amt_max, "Application Deadline": deadline,
        },
    }

def run(max_pages=2, headless=True):
    df = pd.DataFrame(columns=COLUMNS)
    seen_ids, seen_links = set(), set()
    all_entries = []  # (href, source_url)

    # Phase 1: use Playwright only to collect grant links from listing pages
    with sync_playwright() as pw:
        ctx = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        ).new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        )
        lp = ctx.new_page()

        for pn in range(1, max_pages + 1):
            print(f"\n{'='*50}\n  PAGE {pn}/{max_pages}\n{'='*50}")
            source = BASE_URL + f"&page={pn}"
            try:
                lp.goto(source, wait_until="domcontentloaded", timeout=20_000)
                lp.wait_for_selector(
                    "div.tender-item,article.tender-card,.search-results .item,"
                    "[class*='tender'],[class*='opportunity'],.card",
                    timeout=10_000,
                )
            except PwTimeout:
                print("  [WARN] Page load timeout, stopping.")
                break

            entries = get_links(lp)
            if not entries:
                print("  [INFO] No entries found, end of results.")
                break

            for e in entries:
                href = e["href"]
                m = re.search(r"/tenders/(\d+)", href)
                oid = m.group(1) if m else ""
                if href not in seen_links and not (oid and oid in seen_ids):
                    all_entries.append((href, source))
                    seen_links.add(href)
                    if oid:
                        seen_ids.add(oid)

            print(f"  Collected {len(entries)} links ({len(all_entries)} total so far)")

        ctx.browser.close()

    # Phase 2: fetch all detail pages concurrently with requests
    print(f"\nFetching {len(all_entries)} detail pages (10 concurrent)...")

    def fetch(args):
        href, source = args
        return scrape_detail(href, source)

    rows = []
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {pool.submit(fetch, e): e[0] for e in all_entries}
        for i, future in enumerate(as_completed(futures), 1):
            href = futures[future]
            try:
                row = future.result()
            except Exception as e:
                print(f"  [{i:02d}] ERROR {href}: {e}")
                continue
            if row is None:
                print(f"  [{i:02d}] Filtered out: {href[-60:]}")
            else:
                print(f"  [{i:02d}] ✓ {row['Title'][:60]}")
                rows.append(row)

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=COLUMNS)

    # Generate AI summaries in batch for all filtered results
    print(f"\nGenerating AI summaries for {len(df)} opportunities...")
    for idx, row in df.iterrows():
        if row.get("_opp_data"):
            df.at[idx, "AI Summary"] = generate_sam_summary(row["_opp_data"])
    if "_opp_data" in df.columns:
        df = df.drop(columns=["_opp_data"])

    print(f"\n{'='*50}\n  Done — {len(df)} matching rows\n{'='*50}\n")
    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_colwidth", 60)
    pd.set_option("display.width", 220)
    print(df.to_string(index=False))

    SHEET_NAME = "dev_aid"

    if os.path.exists(EXCEL_FILE):
        existing_sheets = pd.ExcelFile(EXCEL_FILE).sheet_names

        if SHEET_NAME in existing_sheets:
            existing_df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

            # Purge expired rows from the existing sheet
            before_purge = len(existing_df)
            existing_df = existing_df[existing_df["Application Deadline"].apply(is_not_expired)]
            purged = before_purge - len(existing_df)
            if purged:
                print(f"Removed {purged} expired grant(s) from existing sheet.")

            # Append only new (not-yet-recorded) grants
            existing_links = set(existing_df["Original Link"])
            new_rows = df[~df["Original Link"].isin(existing_links)]
            combined_df = pd.concat([existing_df, new_rows], ignore_index=True)

            # Rewrite the entire sheet so purged rows are actually gone
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                combined_df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

            if not new_rows.empty:
                print(f"Added {len(new_rows)} new grant(s).")
            else:
                print("No new grants.")

        else:
            # File exists but sheet doesn't — add new sheet without touching others
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
            print(f"Created new sheet '{SHEET_NAME}'")

    else:
        # File doesn't exist at all — create it
        df.to_excel(EXCEL_FILE, sheet_name=SHEET_NAME, index=False)
        print("Created new Excel file")

    apply_impact_formatting(EXCEL_FILE, SHEET_NAME)
    return df

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="DevelopmentAid MENA grant scraper")
    p.add_argument("--pages", type=int, default=2, help="Pages to scrape (default: 2)")
    p.add_argument("--headless", action="store_true", help="Run headless")
    args = p.parse_args()
    print(f"Pages: {args.pages} | Headless: {args.headless}")
    run(args.pages, args.headless)