#!/usr/bin/env python3
"""
EFE MENA Education Grants Scraper
Scrapes Impact Funding Substack for education/workforce grants targeting MENA countries.
"""

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from summarizer import generate_summary
import os
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv

def is_not_expired(deadline_str):
    if not deadline_str:
        return True
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(str(deadline_str).strip(), fmt) >= datetime.today()
        except ValueError:
            continue
    return True

load_dotenv()

OUTPUT_FILE = os.getenv("EXCEL_FILE")

# ── Config ────────────────────────────────────────────────────────────────────

ARCHIVE_URLS = [
    "https://impactfunding.substack.com/s/education-human-rights-and-inclusion/archive?sort=new",
    "https://impactfunding.substack.com/s/health-and-wash/archive?sort=new",
    "https://impactfunding.substack.com/s/agriculture-climate-environment-energy/archive?sort=new",
    "https://impactfunding.substack.com/s/gender/archive?sort=new",
    "https://impactfunding.substack.com/s/humanitarian-aid-emergency-programming/archive?sort=new",
    "https://impactfunding.substack.com/s/cross-cutting-intersectional-impact/archive?sort=new",
    "https://impactfunding.substack.com/s/innovation-research-and-smart-cities/archive?sort=new",
]

#OUTPUT_FILE = "efe_mena_education_grants.xlsx"

MENA_KEYWORDS = [
    "morocco", "algeria", "tunisia", "egypt", "jordan",
    "palestine", "gaza", "west bank", "yemen", "uae",
    "united arab emirates", "saudi arabia", "lebanon",
    "bahrain", "syria", "mena", "middle east", "north africa",
    "arab world", "arab region",
]

TOPIC_KEYWORDS = [
    "youth employment", "workforce development", "employability", "job placement",
    "job creation", "livelihoods", "economic empowerment", "economic inclusion",
    "apprenticeship", "internship", "mentorship", "job readiness", "job search",
    "labor market activation", "economic participation", "labor market entry",
    "neet", "work readiness", "job seekers", "early-career",
    "access to opportunities", "reducing inequalities", "skills development",
    "vocational training", "technical training", "soft skills", "digital skills",
    "vocational skills", "technical skills", "green jobs", "green skills",
    "tvet", "upskilling", "reskilling", "employability skills",
    "curriculum development", "vocational training center", "career center",
    "ai skills", "climate change", "financial literacy", "circular economy",
    "higher education", "university", "educational institutions", "life skills",
    "transversal skills", "entrepreneurial skills", "blended training",
    "entrepreneurship", "sme development", "private sector development",
    "self employment", "virtual jobs", "income generation", "startup incubation",
    "employer engagement", "business acceleration", "micro entrepreneurship",
    "new business creation", "sme", "incubation", "green entrepreneurship",
    "women entrepreneurship", "startup", "startup support", "financial inclusion",
    "home-based businesses", "msme", "microbusiness", "freelance",
    "gig work", "gig economy", "capacity building", "systems strengthening",
    "framework", "action plan", "competitiveness", "skills gaps",
    "business association", "chamber of commerce", "industry federation",
]

def find_chrome_binary():
    candidates = [
        os.getenv("CHROME_BIN"),
        "/usr/bin/google-chrome",
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
    ]

    for pattern in [
        "/ms-playwright/chromium-*/chrome-linux/chrome",
        "/ms-playwright/chromium-*/chrome-linux/headless_shell",
    ]:
        candidates.extend(str(p) for p in Path("/").glob(pattern.lstrip("/")))

    for path in candidates:
        if path and os.path.exists(path):
            return path

    raise FileNotFoundError("Could not find a Chrome/Chromium binary")

# ── Step 1: Collect post URLs ─────────────────────────────────────────────────

def get_post_urls_from_archive(driver, archive_url):
    """Scrape all post URLs from a single archive page by scrolling to the bottom."""
    driver.get(archive_url)

    # Wait for the page itself to load first, not for post links specifically
    WebDriverWait(driver, 30).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )

    last_height = 0
    stable_scrolls = 0

    for _ in range(30):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        new_height = driver.execute_script("return document.body.scrollHeight")

        # If the page stops growing for 2 rounds, stop scrolling
        if new_height == last_height:
            stable_scrolls += 1
        else:
            stable_scrolls = 0

        last_height = new_height

        if stable_scrolls >= 2:
            break

    urls = []
    seen = set()

    for a in driver.find_elements(By.CSS_SELECTOR, "a[href*='/p/']"):
        href = a.get_attribute("href")
        if href:
            clean = href.split("?")[0].split("#")[0]
            if clean not in seen:
                seen.add(clean)
                urls.append(clean)

    print(f"      -> found {len(urls)} raw post link(s)")
    return urls


def get_post_urls():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    #driver = webdriver.Chrome(options=options)
    options = webdriver.ChromeOptions()
    #options.binary_location = find_chrome_binary()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(options=options)
    post_urls = []

    try:
        seen_global = set()
        for archive_url in ARCHIVE_URLS:
            print(f"    Scanning: {archive_url}")
            try:
                urls = get_post_urls_from_archive(driver, archive_url)
            except Exception as e:
                print(f"      [WARN] Failed to scrape archive: {e}")
                continue

            new_count = 0
            for url in urls:
                if url not in seen_global:
                    seen_global.add(url)
                    post_urls.append(url)
                    new_count += 1
            print(f"      -> {new_count} new post(s)")
    finally:
        driver.quit()

    return post_urls


# ── Step 2: Fetch a post ──────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

def fetch_post(url):
    """Return the body markup div, or None if inaccessible."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"    [WARN] {e}")
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    # The actual post content lives in .body.markup inside .available-content
    markup = soup.find("div", class_="body markup")
    if markup:
        return markup

    # Fallback
    return (
        soup.find("div", class_="available-content")
        or soup.find("article")
    )


# ── Step 3: Parse grant blocks ────────────────────────────────────────────────
#
# Actual Substack structure (blockquote + siblings pattern):
#
#   <blockquote>                           ← grant title & donor name
#     <p>
#       <strong>Grant Title</strong>
#       <strong><a href="donor-url">Donor Name</a></strong>
#       <em>*Closing soon!*</em>
#     </p>
#   </blockquote>
#   <p>Description paragraph…</p>          ← sibling
#   <ul>                                   ← sibling
#     <li><p><strong>Geographies:</strong><span> …</span></p></li>
#     <li><p><strong>Who can apply:</strong><span> …</span></p></li>
#     <li><p><strong>Funding amount:</strong><span> …</span></p></li>
#     <li><p><strong>Targeted Sectors / SDGs:</strong><span> …</span></p></li>
#     <li><p><strong>Deadline:</strong><strong> …</strong></p></li>
#     <li><p><strong><a href="…">Learn more and apply here</a></strong></p></li>
#   </ul>
#   <p><em>Italic summary sentence.</em></p>  ← sibling
#   <div><hr/></div>                          ← separator between grants

def assign_field(fields, label, value):
    """Map a label string to the correct field key."""
    label = label.lower().strip()
    if "geograph" in label:
        fields["geographic_area"] = value
    elif "who can apply" in label or "eligib" in label:
        fields["eligibility"] = value
    elif "funding amount" in label or "amount" in label:
        fields["funding_amount"] = value
    elif "sector" in label or "sdg" in label or "focus" in label:
        fields["focus_sector"] = value
    elif "deadline" in label:
        fields["deadline"] = value


def parse_li_fields(ul):
    """
    Extract labeled bullet fields from a grant <ul>.

    Handles three structural variants:
      A) Normal:   <li><p><strong>Label:</strong><span>Value</span></p></li>
      B) Embedded: <li><p><strong>Deadline: May 4, 2025.</strong></p></li>
         (label + value both inside one <strong>)
      C) Multi:    <li><p><strong>Geo:</strong><span>…</span><br/>
                             <strong>Who can apply:</strong><span>…</span></p></li>
         (multiple fields in one <li> separated by <br/>)
    """
    fields = {
        "geographic_area": "",
        "funding_amount": "",
        "focus_sector": "",
        "eligibility": "",
        "deadline": "",
        "grant_link": "",
    }

    for li in ul.find_all("li"):
        p = li.find("p") or li

        # Capture any grant links before touching the tree
        for a in p.find_all("a"):
            if a.get("href"):
                lt = a.get_text(strip=True).lower()
                if "learn" in lt or "apply" in lt or "more" in lt:
                    fields["grant_link"] = a["href"]

        # Iterate over every <strong> that doesn't wrap a link — each one is a field label
        for strong in p.find_all("strong"):
            if strong.find("a"):
                continue

            raw = strong.get_text(strip=True)
            if ":" not in raw:
                continue

            colon = raw.index(":")
            label = raw[:colon].strip()
            value_in_strong = raw[colon + 1:].strip()

            # Collect additional value text from siblings.
            # Stop at <br/> (next field in same <li>) or a <strong> that looks like
            # another label (has a colon within its first ~30 chars).
            # Value <strong>s (e.g. <strong>March 3, 2026</strong>) are included.
            extra = []
            for sib in strong.next_siblings:
                sib_name = getattr(sib, "name", None)
                if sib_name == "br":
                    break
                if sib_name == "strong":
                    sib_text = sib.get_text(strip=True)
                    colon_pos = sib_text.find(":")
                    if colon_pos != -1 and colon_pos < 35:
                        break  # this is another label — stop
                text = sib.get_text(strip=True) if hasattr(sib, "get_text") else str(sib).strip()
                if text:
                    extra.append(text)

            value = (value_in_strong + " " + " ".join(extra)).strip().lstrip(":").strip()
            assign_field(fields, label, value)

    return fields


def parse_grant_group(bq, siblings):
    """
    Build a grant dict from a <blockquote> (title/donor) and its
    following sibling tags (<p> description, <ul> bullets, <p><em> summary).
    """
    # Title & donor from blockquote
    strongs = bq.find_all("strong")
    title = strongs[0].get_text(strip=True) if strongs else ""

    donor_name = ""
    for s in strongs:
        a = s.find("a")
        if a:
            donor_name = a.get_text(strip=True)
            break

    # Fallback: if title+donor are merged in one <strong> (no link),
    # split on the last comma: "Grant Title, Donor Name"
    if not donor_name and "," in title:
        parts = title.rsplit(",", 1)
        title = parts[0].strip()
        donor_name = parts[1].strip()

    description = ""
    fields = {
        "geographic_area": "",
        "funding_amount": "",
        "focus_sector": "",
        "eligibility": "",
        "deadline": "",
        "grant_link": "",
    }

    for tag in siblings:
        if tag.name == "ul":
            fields = parse_li_fields(tag)
        elif tag.name == "p" and not description:
            # Use the first plain paragraph (above the bullet list) as the summary.
            # Skip paragraphs that are only italic markers like "*Closing soon!*"
            text = tag.get_text(strip=True)
            if text and not text.startswith("*"):
                description = text

    return {
        "title": title,
        "donor_name": donor_name,
        "description": description,
        "summary": "",
        **fields,
    }


def is_relevant(grant):
    """True if the grant covers a MENA country/region AND matches a topic keyword."""
    haystack = " ".join([
        grant.get("geographic_area", ""),
        grant.get("title", ""),
        grant.get("description", ""),
        grant.get("eligibility", ""),
        grant.get("focus_sector", ""),
    ]).lower()

    has_mena = any(kw in haystack for kw in MENA_KEYWORDS)
    has_topic = any(kw in haystack for kw in TOPIC_KEYWORDS)
    return has_mena and has_topic


def parse_grants_from_post(markup, post_url):
    """
    Walk the top-level children of the markup div.
    Each <blockquote> starts a new grant; collect siblings until the next
    blockquote or <hr> separator.
    """
    grants = []
    children = [c for c in markup.children if c.name]

    i = 0
    while i < len(children):
        tag = children[i]
        if tag.name == "blockquote":
            # Collect following siblings until next grant or separator
            siblings = []
            j = i + 1
            while j < len(children):
                sib = children[j]
                if sib.name == "blockquote":
                    break
                if sib.name == "div" and sib.find("hr"):
                    j += 1  # consume the <hr> div
                    break
                siblings.append(sib)
                j += 1

            grant = parse_grant_group(tag, siblings)
            if is_relevant(grant):
                grant["post_link"] = post_url
                grants.append(grant)

            i = j
        else:
            i += 1

    return grants


# ── Step 4: Save to Excel ─────────────────────────────────────────────────────

COLUMNS = [
    ("Post Link",                50),
    ("Donor Name",               28),
    ("Funding Amount (USD)",     22),
    ("Geographic Area",          22),
    ("Focus / Sector",           28),
    ("Eligibility Requirements", 45),
    ("Grant Link",               45),
    ("Deadline",                 18),
    ("Description",              65),
    ("AI Summary",               65),
]

GRANT_KEYS = [
    "post_link", "donor_name", "funding_amount",
    "geographic_area", "focus_sector", "eligibility",
    "grant_link", "deadline", "description", "summary",
]

HEADER_COLOR = "1F4E79"

def save_to_excel(grants, path=OUTPUT_FILE):
    sheet_name = "impact funding"

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        # Load existing rows and purge expired ones
        if sheet_name in wb.sheetnames:
            import pandas as pd
            existing_df = pd.read_excel(path, sheet_name=sheet_name)
            before_purge = len(existing_df)
            existing_df = existing_df[existing_df["Deadline"].apply(is_not_expired)]
            purged = before_purge - len(existing_df)
            if purged:
                print(f"  Removed {purged} expired grant(s) from existing sheet.")
            existing_links = set(existing_df["Grant Link"].dropna())
        else:
            existing_df = None
            existing_links = set()
    else:
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name  # placeholder so it exists
        existing_df = None
        existing_links = set()

    # Deduplicate incoming grants against existing
    new_grants = [g for g in grants if g.get("grant_link", "") not in existing_links]
    print(f"  {len(new_grants)} new grant(s) to add.")

    # Rebuild sheet from scratch
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    wrap_top = Alignment(vertical="top", wrap_text=True)
    row_idx = 2

    # Write back surviving existing rows
    if existing_df is not None:
        for _, row in existing_df.iterrows():
            for col_idx, (header, _) in enumerate(COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, "")).alignment = wrap_top
            ws.row_dimensions[row_idx].height = 80
            row_idx += 1

    # Append new grants
    for grant in new_grants:
        for col_idx, key in enumerate(GRANT_KEYS, 1):
            ws.cell(row=row_idx, column=col_idx, value=grant.get(key, "")).alignment = wrap_top
        ws.row_dimensions[row_idx].height = 80
        row_idx += 1

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  Saved {row_idx - 2} total grants -> {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  EFE MENA Grants Scraper")
    print("=" * 60)

    print(f"\n[1/4] Collecting post URLs from {len(ARCHIVE_URLS)} archive(s)...")
    post_urls = get_post_urls()
    print(f"  Found {len(post_urls)} posts")

    print("\n[2/4] Fetching posts and parsing grants...\n")
    all_grants = []

    for i, url in enumerate(post_urls, 1):
        print(f"  [{i:02d}/{len(post_urls)}] {url}")

        markup = fetch_post(url)
        if markup is None:
            print("    -> Skipped (no accessible content — likely paywalled)")
            time.sleep(1)
            continue

        grants = parse_grants_from_post(markup, url)
        print(f"    -> {len(grants)} qualifying grant(s) found")
        all_grants.extend(grants)
        time.sleep(1)

    print(f"\n[3/4] Generating AI summaries for {len(all_grants)} grant(s)...")
    for i, grant in enumerate(all_grants):
        #grant["summary"] = "test"
        grant["summary"] = generate_summary(grant)
        print(f"    [{i+1}/{len(all_grants)}] summarized")

    print(f"\n[4/4] Exporting {len(all_grants)} total grant(s) to Excel...")
    all_grants = [g for g in all_grants if is_not_expired(g.get("deadline", ""))]
    print(f"  After deadline filter: {len(all_grants)} grant(s)")
    save_to_excel(all_grants)
    print("\nDone!")


if __name__ == "__main__":
    main()
