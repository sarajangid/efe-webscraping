#!/usr/bin/env python3
"""
EFE MENA Education Grants Scraper
Scrapes Impact Funding Substack for education grants targeting MENA countries.
"""

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from summarizer import generate_summary

# ── Config ────────────────────────────────────────────────────────────────────

ARCHIVE_URL = (
    "https://impactfunding.substack.com/s/education-human-rights-and-inclusion"
    "/archive?sort=new"
)

OUTPUT_FILE = "efe_mena_education_grants.xlsx"

MENA_KEYWORDS = [
    "morocco", "algeria", "tunisia", "egypt", "jordan",
    "palestine", "gaza", "west bank", "yemen", "uae",
    "united arab emirates", "saudi arabia", "lebanon",
    "bahrain", "syria", "mena", "middle east", "north africa",
    "arab world", "arab region",
]

# ── Step 1: Collect post URLs ─────────────────────────────────────────────────

def get_post_urls():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(options=options)
    post_urls = []

    try:
        driver.get(ARCHIVE_URL)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/p/']"))
        )

        last_height = 0
        for _ in range(20):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        seen = set()
        for a in driver.find_elements(By.CSS_SELECTOR, "a[href*='/p/']"):
            href = a.get_attribute("href")
            if href:
                clean = href.split("?")[0].split("#")[0]
                if clean not in seen:
                    seen.add(clean)
                    post_urls.append(clean)
    finally:
        driver.quit()

    return post_urls


# ── Step 2: Fetch a post ──────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

def fetch_post(url):
    """Return the body markup div, or None if inaccessible."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"    [WARN] {e}")
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    # The actual post content lives in .body.markup inside .available-content
    markup = soup.find("div", class_="body markup")
    if markup:
        return markup

    # Fallback
    return (
        soup.find("div", class_="available-content")
        or soup.find("article")
    )


# ── Step 3: Parse grant blocks ────────────────────────────────────────────────
#
# Actual Substack structure (blockquote + siblings pattern):
#
#   <blockquote>                           ← grant title & donor name
#     <p>
#       <strong>Grant Title</strong>
#       <strong><a href="donor-url">Donor Name</a></strong>
#       <em>*Closing soon!*</em>
#     </p>
#   </blockquote>
#   <p>Description paragraph…</p>          ← sibling
#   <ul>                                   ← sibling
#     <li><p><strong>Geographies:</strong><span> …</span></p></li>
#     <li><p><strong>Who can apply:</strong><span> …</span></p></li>
#     <li><p><strong>Funding amount:</strong><span> …</span></p></li>
#     <li><p><strong>Targeted Sectors / SDGs:</strong><span> …</span></p></li>
#     <li><p><strong>Deadline:</strong><strong> …</strong></p></li>
#     <li><p><strong><a href="…">Learn more and apply here</a></strong></p></li>
#   </ul>
#   <p><em>Italic summary sentence.</em></p>  ← sibling
#   <div><hr/></div>                          ← separator between grants

def assign_field(fields, label, value):
    """Map a label string to the correct field key."""
    label = label.lower().strip()
    if "geograph" in label:
        fields["geographic_area"] = value
    elif "who can apply" in label or "eligib" in label:
        fields["eligibility"] = value
    elif "funding amount" in label or "amount" in label:
        fields["funding_amount"] = value
    elif "sector" in label or "sdg" in label or "focus" in label:
        fields["focus_sector"] = value
    elif "deadline" in label:
        fields["deadline"] = value


def parse_li_fields(ul):
    """
    Extract labeled bullet fields from a grant <ul>.

    Handles three structural variants:
      A) Normal:   <li><p><strong>Label:</strong><span>Value</span></p></li>
      B) Embedded: <li><p><strong>Deadline: May 4, 2025.</strong></p></li>
         (label + value both inside one <strong>)
      C) Multi:    <li><p><strong>Geo:</strong><span>…</span><br/>
                             <strong>Who can apply:</strong><span>…</span></p></li>
         (multiple fields in one <li> separated by <br/>)
    """
    fields = {
        "geographic_area": "",
        "funding_amount": "",
        "focus_sector": "",
        "eligibility": "",
        "deadline": "",
        "grant_link": "",
    }

    for li in ul.find_all("li"):
        p = li.find("p") or li

        # Capture any grant links before touching the tree
        for a in p.find_all("a"):
            if a.get("href"):
                lt = a.get_text(strip=True).lower()
                if "learn" in lt or "apply" in lt or "more" in lt:
                    fields["grant_link"] = a["href"]

        # Iterate over every <strong> that doesn't wrap a link — each one is a field label
        for strong in p.find_all("strong"):
            if strong.find("a"):
                continue

            raw = strong.get_text(strip=True)
            if ":" not in raw:
                continue

            colon = raw.index(":")
            label = raw[:colon].strip()
            value_in_strong = raw[colon + 1:].strip()

            # Collect additional value text from siblings.
            # Stop at <br/> (next field in same <li>) or a <strong> that looks like
            # another label (has a colon within its first ~30 chars).
            # Value <strong>s (e.g. <strong>March 3, 2026</strong>) are included.
            extra = []
            for sib in strong.next_siblings:
                sib_name = getattr(sib, "name", None)
                if sib_name == "br":
                    break
                if sib_name == "strong":
                    sib_text = sib.get_text(strip=True)
                    colon_pos = sib_text.find(":")
                    if colon_pos != -1 and colon_pos < 35:
                        break  # this is another label — stop
                text = sib.get_text(strip=True) if hasattr(sib, "get_text") else str(sib).strip()
                if text:
                    extra.append(text)

            value = (value_in_strong + " " + " ".join(extra)).strip().lstrip(":").strip()
            assign_field(fields, label, value)

    return fields


def parse_grant_group(bq, siblings):
    """
    Build a grant dict from a <blockquote> (title/donor) and its
    following sibling tags (<p> description, <ul> bullets, <p><em> summary).
    """
    # Title & donor from blockquote
    strongs = bq.find_all("strong")
    title = strongs[0].get_text(strip=True) if strongs else ""

    donor_name = ""
    for s in strongs:
        a = s.find("a")
        if a:
            donor_name = a.get_text(strip=True)
            break

    # Fallback: if title+donor are merged in one <strong> (no link),
    # split on the last comma: "Grant Title, Donor Name"
    if not donor_name and "," in title:
        parts = title.rsplit(",", 1)
        title = parts[0].strip()
        donor_name = parts[1].strip()

    description = ""
    fields = {
        "geographic_area": "",
        "funding_amount": "",
        "focus_sector": "",
        "eligibility": "",
        "deadline": "",
        "grant_link": "",
    }

    for tag in siblings:
        if tag.name == "ul":
            fields = parse_li_fields(tag)
        elif tag.name == "p" and not description:
            # Use the first plain paragraph (above the bullet list) as the summary.
            # Skip paragraphs that are only italic markers like "*Closing soon!*"
            text = tag.get_text(strip=True)
            if text and not text.startswith("*"):
                description = text

    return {
        "title": title,
        "donor_name": donor_name,
        "description": description,
        "summary": "",
        **fields,
    }


def is_relevant(grant):
    """True if the grant covers any of the target MENA countries/regions."""
    haystack = " ".join([
        grant.get("geographic_area", ""),
        grant.get("title", ""),
        grant.get("summary", ""),
        grant.get("eligibility", ""),
    ]).lower()

    return any(kw in haystack for kw in MENA_KEYWORDS)


def parse_grants_from_post(markup, post_url):
    """
    Walk the top-level children of the markup div.
    Each <blockquote> starts a new grant; collect siblings until the next
    blockquote or <hr> separator.
    """
    grants = []
    children = [c for c in markup.children if c.name]

    i = 0
    while i < len(children):
        tag = children[i]
        if tag.name == "blockquote":
            # Collect following siblings until next grant or separator
            siblings = []
            j = i + 1
            while j < len(children):
                sib = children[j]
                if sib.name == "blockquote":
                    break
                if sib.name == "div" and sib.find("hr"):
                    j += 1  # consume the <hr> div
                    break
                siblings.append(sib)
                j += 1

            grant = parse_grant_group(tag, siblings)
            if is_relevant(grant):
                grant["post_link"] = post_url
                grants.append(grant)

            i = j
        else:
            i += 1

    return grants


# ── Step 4: Save to Excel ─────────────────────────────────────────────────────

COLUMNS = [
    ("Post Link",                50),
    ("Donor Name",               28),
    ("Funding Amount (USD)",     22),
    ("Geographic Area",          22),
    ("Focus / Sector",           28),
    ("Eligibility Requirements", 45),
    ("Grant Link",               45),
    ("Deadline",                 18),
    ("Description",              65),
    ("AI Summary",               65),
]

GRANT_KEYS = [
    "post_link", "donor_name", "funding_amount",
    "geographic_area", "focus_sector", "eligibility",
    "grant_link", "deadline", "description", "summary",
]

HEADER_COLOR = "1F4E79"

def save_to_excel(grants, path=OUTPUT_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MENA Education Grants"

    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    wrap_top = Alignment(vertical="top", wrap_text=True)

    for row_idx, grant in enumerate(grants, 2):
        for col_idx, key in enumerate(GRANT_KEYS, 1):
            ws.cell(row=row_idx, column=col_idx, value=grant.get(key, "")).alignment = wrap_top
        ws.row_dimensions[row_idx].height = 80

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"\n  Saved {len(grants)} grants -> {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  EFE MENA Education Grants Scraper")
    print("=" * 60)

    print("\n[1/3] Collecting post URLs from archive...")
    post_urls = get_post_urls()
    print(f"  Found {len(post_urls)} posts")

    print("\n[2/3] Fetching posts and parsing grants...\n")
    all_grants = []

    for i, url in enumerate(post_urls, 1):
        print(f"  [{i:02d}/{len(post_urls)}] {url}")

        markup = fetch_post(url)
        if markup is None:
            print("    -> Skipped (no accessible content — likely paywalled)")
            time.sleep(1)
            continue

        grants = parse_grants_from_post(markup, url)
        print(f"    -> {len(grants)} qualifying grant(s) found")
        all_grants.extend(grants)
        time.sleep(1)

    print(f"\n[3/4] Generating AI summaries for {len(all_grants)} grant(s)...")
    for i, grant in enumerate(all_grants):
        grant["summary"] = generate_summary(grant)
        print(f"    [{i+1}/{len(all_grants)}] summarized")
        if i < len(all_grants) - 1:
            time.sleep(13)  # stay under 5 req/min free tier limit

    print(f"\n[4/4] Exporting {len(all_grants)} total grant(s) to Excel...")
    save_to_excel(all_grants)
    print("\nDone!")


if __name__ == "__main__":
    main()
