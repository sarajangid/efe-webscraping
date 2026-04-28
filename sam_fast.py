import argparse, re, urllib.parse, asyncio, time
from datetime import datetime
import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PwTimeout
import os
from dotenv import load_dotenv
from summarizer import generate_sam_summary
from reqs import MENA_COUNTRIES, KEYWORDS, COLUMNS
from dev_aid import norm, matches, parse_amount
from concurrent.futures import ThreadPoolExecutor

def is_not_expired(deadline_str):
    if not deadline_str:
        return True
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(str(deadline_str).strip(), fmt) >= datetime.today()
        except ValueError:
            continue
    return True

def ts() -> str:
    return datetime.now().strftime("[%H:%M:%S]")

load_dotenv()
EXCEL_FILE = os.environ["EXCEL_FILE"]

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_COLOR = "1F4E79"
HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

DEFAULT_WIDTHS = {
    "Opportunity ID": 18,
    "Opportunity Type": 22,
    "Title": 40,
    "Donor Name": 28,
    "Geographic Area": 24,
    "Focus / Sector": 28,
    "Application Deadline": 20,
    "Amount Min (USD)": 18,
    "Amount Max (USD)": 18,
    "Eligibility": 42,
    "Eligibility Requirements": 45,
    "Matched Keywords": 32,
    "Source Link": 45,
    "Original Link": 45,
    "Date Posted": 18,
    "Date Scraped": 16,
    "Grant Name": 38,
    "Agency": 28,
    "Due Date": 18,
    "Award Minimum": 18,
    "Award Maximum": 18,
    "Description": 65,
    "Documents": 45,
    "Application Link": 45,
    "AI Summary": 65,
    "Post Link": 50,
    "Funding Amount (USD)": 22,
    "Grant Link": 45,
    "Deadline": 18,
}

def _auto_width(header: str) -> float:
    header = (header or "").strip()
    if not header:
        return 18
    return min(max(len(header) + 4, 18), 45)

def apply_impact_formatting(path: str, sheet_name: str, custom_widths=None):
    wb = openpyxl.load_workbook(path)

    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    widths = {**DEFAULT_WIDTHS, **(custom_widths or {})}

    if ws.max_row >= 1:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            header = "" if cell.value is None else str(cell.value)

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(
                header, _auto_width(header)
            )

        ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = BODY_ALIGNMENT

    ws.freeze_panes = "A2"
    wb.save(path)

BASE_URL = "https://sam.gov/search/?index=_all&sort=-modifiedDate&sfm%5Bstatus%5D%5Bis_active%5D=true&sfm%5Bstatus%5D%5Bis_inactive%5D=true&sfm%5BsimpleSearch%5D%5BkeywordRadio%5D=ALL"

# How many detail pages to scrape in parallel
CONCURRENCY = 5

SECTOR_MAP = {
    "Youth Workforce Development": {"youth employment","job placement","job readiness","neet","work readiness","early-career","job seekers"},
    "Entrepreneurship / SME Development": {"entrepreneurship","sme","sme development","msme","startup","startup support","micro entrepreneurship","microbusiness","women entrepreneurship","green entrepreneurship","startup incubation","business acceleration"},
    "Vocational Training / Skills": {"vocational training","tvet","skills development","upskilling","reskilling","technical training","blended training","curriculum development","employability skills"},
    "Green Economy": {"green jobs","green skills","green entrepreneurship","circular economy"},
    "Financial Inclusion / Literacy": {"financial literacy","financial inclusion"},
    "Digital Skills": {"digital skills"},
    "Economic Empowerment / Livelihoods": {"economic empowerment","economic inclusion","income generation","livelihoods","self employment","gig economy","freelance"},
    "Capacity Building / Systems Strengthening": {"capacity building","systems strengthening","competitiveness","skills gaps","business association","chamber of commerce","industry federation","private sector development"},
}

def infer_sector(matched):
    return " | ".join(s for s, terms in SECTOR_MAP.items() if {k.lower() for k in matched} & terms) or "Workforce / Economic Development"

def parse_eligibility(text):
    patterns = [r'(?:eligible|eligibility)[^.]{0,200}\.', r'(?:open to|restricted to|available to)\s+[^.]{5,200}\.', r'(?:only|exclusively)\s+(?:available|open)\s+(?:to|for)\s+[^.]{5,150}\.', r'(?:non-?profit|NGO|INGO|government entity|private sector)[^.]{0,150}(?:eligible|may apply|can apply)[^.]*\.', r'applicants?\s+must\s+be[^.]{5,200}\.']
    seen, hits = set(), []
    [hits.append(m.strip()) or seen.add(m.strip()) for pat in patterns for m in re.findall(pat, text or "", re.IGNORECASE) if m.strip() not in seen]
    return " | ".join(hits[:3])


async def aget(page, sel):
    """Async equivalent of dev_aid.get()."""
    el = await page.query_selector(sel)
    return norm(await el.inner_text()) if el else ""


async def get_links_async(page, keyword, pg, skip_ids):
    """Fetch one listing page and return new opp URLs not in skip_ids."""
    try:
        await page.goto(
            f"{BASE_URL}&page={pg}&pageSize=25&sfm%5BsimpleSearch%5D%5BkeywordTags%5D%5B0%5D%5Bvalue%5D={urllib.parse.quote(keyword)}",
            wait_until="domcontentloaded", timeout=30_000,
        )
        await page.wait_for_selector("a[href*='/opp/']", timeout=10_000)
    except PwTimeout:
        return []
    # Extract all hrefs in one JS call — faster than per-element awaits
    hrefs = await page.eval_on_selector_all(
        "a[href*='/opp/']",
        "els => [...new Set(els.map(e => e.href))]",
    )
    print(f"{ts()}     [debug] page {pg} | hrefs found: {len(hrefs)}")
    return [h for h in hrefs if (m := re.search(r"/opp/([^/]+)/", h)) and m.group(1) not in skip_ids]


async def scrape_detail_async(semaphore, ctx, opp_url):
    """Scrape one detail page inside the concurrency semaphore; each call owns its own page."""
    async with semaphore:
        page = await ctx.new_page()
        try:
            if not (m := re.search(r"/opp/([^/]+)/", opp_url)):
                return None
            opp_id = m.group(1)
            try:
                await page.goto(opp_url, wait_until="domcontentloaded", timeout=30_000)
                await page.wait_for_selector(
                    "[class*='opportunity-title'],[class*='opp-title'],h1", timeout=15_000
                )
                # Removed wait_for_timeout(1500) — selector wait above is sufficient
                body = norm(await page.inner_text("body"))

                opp_type = await aget(page, "[class*='notice-type'],[class*='opportunity-type'],[class*='noticeType'],.type-label")
                if opp_type and not any(a in opp_type.lower() for a in {"solicitation","presolicitation","sources sought","special notice","grant","cooperative agreement"}):
                    print(f"{ts()}     [filtered] opp_type='{opp_type}' | {opp_id}"); return None

                mena    = matches(body, MENA_COUNTRIES)
                if not mena:    print(f"{ts()}     [filtered] no MENA | {opp_id}"); return None
                matched = matches(body, KEYWORDS)
                if not matched: print(f"{ts()}     [filtered] no keywords | mena={mena} | {opp_id}"); return None

                title    = await aget(page, "h1,[class*='opportunity-title'],[class*='opp-title']")
                donor    = await aget(page, "[class*='organization'],[class*='agency-name'],[class*='dept']")
                deadline = await aget(page, "[class*='deadline'],[class*='response-date'],[class*='responseDeadline'],.response-deadline")
                amt      = parse_amount(await aget(page, "[class*='award-amount'],[class*='amount']"))
                # Extract all external hrefs in one JS call
                external = await page.eval_on_selector_all(
                    "a[href^='http']",
                    "els => els.map(e => e.href).filter(h => !h.includes('sam.gov'))",
                )
                date_posted = await aget(page, "[class*='posted-date'],[class*='postDate']")
                opp_data = {"Title": title, "Donor Name": donor, "Geographic Area": ", ".join(mena), "Focus / Sector": infer_sector(matched), "Eligibility": parse_eligibility(body), "Amount Max (USD)": amt[1], "Application Deadline": deadline}

                print(f"{ts()}     ✓ [{opp_id}] '{title[:55]}' | KW: {' | '.join(matched)[:60]}")
                return {"Opportunity ID": opp_id, "Opportunity Type": opp_type, "Title": title, "Donor Name": donor, "Geographic Area": ", ".join(mena), "Focus / Sector": infer_sector(matched), "Application Deadline": deadline, "Amount Min (USD)": amt[0], "Amount Max (USD)": amt[1], "Eligibility": parse_eligibility(body), "Matched Keywords": " | ".join(matched), "Source Link": opp_url, "Original Link": external[0] if external else "", "Date Posted": date_posted, "AI Summary": None, "_opp_data": opp_data}
            except PwTimeout:
                print(f"{ts()}     [WARN] Timeout: {opp_url}"); return None
            except Exception as exc:
                print(f"{ts()}     Error on {opp_url}: {exc}"); return None
        finally:
            await page.close()


async def _run_async(max_pages, headless):
    # Pre-load existing IDs from Excel so we skip re-scraping them entirely
    skip_ids: set[str] = set()
    if os.path.exists(EXCEL_FILE):
        try:
            ef = pd.ExcelFile(EXCEL_FILE)
            if "sam" in ef.sheet_names:
                skip_ids = set(pd.read_excel(EXCEL_FILE, sheet_name="sam")["Opportunity ID"].astype(str))
                print(f"{ts()} Pre-loaded {len(skip_ids)} existing IDs from Excel — these will be skipped.")
        except Exception as e:
            print(f"{ts()} [WARN] Could not pre-load existing IDs: {e}")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=headless, args=["--disable-blink-features=AutomationControlled"]
        )
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        )

        # ── Phase 1: Collect all unique URLs across all keywords ──────────────
        # Single listing page, sequential — avoids bot detection.
        # Early-stop per keyword the moment a page returns no new links.
        lp = await ctx.new_page()
        all_urls: dict[str, str] = {}  # opp_id -> url

        for i, keyword in enumerate(KEYWORDS, 1):
            print(f"\n{'='*50}\n{ts()}   KEYWORD [{i:3}/{len(KEYWORDS)}] '{keyword}'\n{'='*50}")
            for pg in range(1, max_pages + 1):
                combined_skip = skip_ids | set(all_urls.keys())
                new_hrefs = await get_links_async(lp, keyword, pg, combined_skip)
                if not new_hrefs:
                    print(f"{ts()}     [INFO] No new links on page {pg} — stopping this keyword early.")
                    break
                for href in new_hrefs:
                    if m := re.search(r"/opp/([^/]+)/", href):
                        all_urls[m.group(1)] = href
                print(f"{ts()}     Unique new URLs queued: {len(all_urls)}")

        await lp.close()

        # ── Phase 2: Scrape all detail pages concurrently ────────────────────
        print(f"\n{'='*50}\n{ts()}   Phase 2: Scraping {len(all_urls)} unique URLs (concurrency={CONCURRENCY})\n{'='*50}")
        semaphore = asyncio.Semaphore(CONCURRENCY)
        tasks = [scrape_detail_async(semaphore, ctx, url) for url in all_urls.values()]
        results = await asyncio.gather(*tasks)

        await browser.close()

    rows = [r for r in results if r is not None]

    # ── Phase 3: AI summaries — run concurrently via threads ─────────────────
    print(f"\n{ts()} Generating AI summaries for {len(rows)} opportunities...")
    def _gen(row):
        return generate_sam_summary(row["_opp_data"]) if row.get("_opp_data") else ""

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [executor.submit(_gen, row) for row in rows]
    # futures are all done by the time we exit the `with` block
    for row, fut in zip(rows, futures):
        row["AI Summary"] = fut.result()

    for row in rows:
        row.pop("_opp_data", None)

    # Build DataFrame once from the collected list — avoids O(n²) pd.concat loop
    return pd.DataFrame(rows, columns=COLUMNS) if rows else pd.DataFrame(columns=COLUMNS)


def run(max_pages=2, headless=True):
    start = time.time()
    df = asyncio.run(_run_async(max_pages, headless))

    print(f"\n{'='*50}\n{ts()}   Done — {len(df)} matching rows\n{'='*50}\n")
    pd.set_option("display.max_columns", None); pd.set_option("display.max_colwidth", 60); pd.set_option("display.width", 220)
    print(df.to_string(index=False))
    
    df = df[df["Application Deadline"].apply(is_not_expired)]
    print(f"{ts()} After deadline filter: {len(df)} rows remaining")

    SHEET_NAME = "sam"
    if os.path.exists(EXCEL_FILE):
        if SHEET_NAME in pd.ExcelFile(EXCEL_FILE).sheet_names:
            new_rows = df[~df["Opportunity ID"].isin(set(pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)["Opportunity ID"].astype(str)))]
            if not new_rows.empty:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    new_rows.to_excel(writer, sheet_name=SHEET_NAME, startrow=writer.book[SHEET_NAME].max_row, index=False, header=False)
                print(f"{ts()} Added {len(new_rows)} new grants")
            else: print(f"{ts()} No new grants")
        else:
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer: df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
            print(f"{ts()} Created new sheet '{SHEET_NAME}'")
    else:
        df.to_excel(EXCEL_FILE, sheet_name=SHEET_NAME, index=False); print(f"{ts()} Created new Excel file")
    apply_impact_formatting(EXCEL_FILE, SHEET_NAME)

    elapsed = time.time() - start
    mins, secs = divmod(int(elapsed), 60)
    print(f"\n{ts()} Total time: {mins}m {secs}s")
    return df


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="SAM.gov MENA grant scraper")
    p.add_argument("--pages", type=int, default=2, help="Pages per keyword (default: 2)")
    p.add_argument("--headless", action="store_true", help="Run headless")
    args = p.parse_args()
    print(f"{ts()} Pages: {args.pages} | Headless: {args.headless}")
    run(args.pages, args.headless)
