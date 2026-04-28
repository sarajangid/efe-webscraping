import datetime
import os
import re
import json
import time
import shutil
import requests
import pandas as pd
from urllib.parse import urlencode
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException

from upload_to_sharepoint import download_documents
from summarizer import generate_simpler_summary

load_dotenv()
EXCEL_FILE=os.environ["EXCEL_FILE"]
BASE_DOWNLOAD_DIR = os.environ["BASE_DOWNLOAD_DIR"]

MAX_SEARCH_PAGES = int(os.getenv("SIMPLER_MAX_PAGES", "900"))
DETAIL_REQUEST_TIMEOUT = 30

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_COLOR = "1F4E79"
HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

DEFAULT_WIDTHS = {
    "Opportunity ID": 18,
    "Opportunity Type": 22,
    "Title": 40,
    "Donor Name": 28,
    "Geographic Area": 24,
    "Focus / Sector": 28,
    "Application Deadline": 20,
    "Amount Min (USD)": 18,
    "Amount Max (USD)": 18,
    "Eligibility": 42,
    "Eligibility Requirements": 45,
    "Matched Keywords": 32,
    "Source Link": 45,
    "Original Link": 45,
    "Date Posted": 18,
    "Date Scraped": 16,
    "Grant Name": 38,
    "Agency": 28,
    "Due Date": 18,
    "Award Minimum": 18,
    "Award Maximum": 18,
    "Description": 65,
    "Documents": 45,
    "Application Link": 45,
    "AI Summary": 65,
    "Post Link": 50,
    "Funding Amount (USD)": 22,
    "Grant Link": 45,
    "Deadline": 18,
}

ALLOWED_AGENCIES = {
    "Bureau of Africa Regional Services",
    "Assistance Coordination",
    "Bureau of African Affairs",
    "Bureau of Democracy Human Rights and Labor",
    "Bureau of Economic and Business Affairs",
    "Office of Global Women's Issues",
    "Office of the Middle East Partnership Initiative",
    "Bureau of Educational and Cultural Affairs",
    "Bureau of Population Refugees and Migration",
    "Bureau of Disaster and Humanitarian Response",
    "Bureau of Near Eastern Affairs",
    "Bureau of Global Public Affairs",
    "Iraq Assistance Office",
    "US Mission to Algeria",
    "US Mission to Bahrain",
    "US Mission to Egypt",
    "US Mission to Iraq",
    "US Mission to Israel",
    "US Mission to Jerusalem",
    "US Mission to Jordan",
    "US Mission to Kuwait",
    "US Mission to Lebanon",
    "US Mission to Libya",
    "US Mission to Mauritania",
    "US Mission to Morocco",
    "US Mission to Oman",
    "US Mission to Qatar",
    "US Mission to Saudi Arabia",
    "US Mission to Tunisia",
    "US Mission to United Arab Emirates",
    "Bureau of International Labor Affairs",
    "Millennium Challenge Corporation",
}

def is_not_expired(deadline_str):
    if not deadline_str:
        return True
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%b %d, %Y"):
        try:
            return datetime.datetime.strptime(str(deadline_str).strip(), fmt) >= datetime.datetime.today()
        except ValueError:
            continue
    return True

def _auto_width(header: str) -> float:
    header = (header or "").strip()
    if not header:
        return 18
    return min(max(len(header) + 4, 18), 45)

def apply_impact_formatting(path: str, sheet_name: str, custom_widths=None):
    wb = openpyxl.load_workbook(path)

    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    widths = {**DEFAULT_WIDTHS, **(custom_widths or {})}

    if ws.max_row >= 1:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            header = "" if cell.value is None else str(cell.value)

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(
                header, _auto_width(header)
            )

        ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = BODY_ALIGNMENT

    ws.freeze_panes = "A2"
    wb.save(path)

############################
# CONFIG
############################

def find_chrome_binary():
    candidates = [
        os.getenv("CHROME_BIN"),
        "/usr/bin/google-chrome",
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
    ]

    for pattern in [
        "/ms-playwright/chromium-*/chrome-linux/chrome",
        "/ms-playwright/chromium-*/chrome-linux/headless_shell",
    ]:
        candidates.extend(str(p) for p in Path("/").glob(pattern.lstrip("/")))

    for path in candidates:
        if path and os.path.exists(path):
            return path

    raise FileNotFoundError("Could not find a Chrome/Chromium binary")

BASE_SEARCH_URL = "https://simpler.grants.gov/search"
BASE_DOMAIN = "https://simpler.grants.gov"

params = {
    "andOr": "OR",
    "query": "education science technology engineering math career",
}

SHEET_NAME = "SimplerGrants"

os.makedirs(os.path.join(BASE_DOWNLOAD_DIR,SHEET_NAME), exist_ok=True)


############################
# COLLECT SEARCH LINKS
############################

search_url = BASE_SEARCH_URL + "?" + urlencode(params)

options = webdriver.ChromeOptions()
#options.binary_location = find_chrome_binary()
options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--disable-blink-features=AutomationControlled")

driver = webdriver.Chrome(options=options)

#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get(search_url)

wait = WebDriverWait(driver, 10)

rows = []
links = []

for page_num in range(1, MAX_SEARCH_PAGES + 1):
    print(f"Scanning search page {page_num}/{MAX_SEARCH_PAGES}")

    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tr.border-base")))
    except TimeoutException:
        print("Timed out waiting for search results table; stopping pagination.")
        break

    soup = BeautifulSoup(driver.page_source, "html.parser")
    results = soup.select("tr.border-base")

    if not results:
        print("No search results found on page; stopping pagination.")
        break

    before_count = len(links)

    for result in results:
        a_tag = result.select_one("a")
        if not a_tag:
            continue

        relative_link = a_tag.get("href")
        if not relative_link:
            continue

        full_link = BASE_DOMAIN + relative_link
        if full_link not in links:
            links.append(full_link)

    print(f"Collected {len(links)} total links (+{len(links) - before_count} new)")

    if page_num >= MAX_SEARCH_PAGES:
        print("Reached max search pages limit.")
        break

    try:
        next_button = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "button[data-testid='pagination-next']")
            )
        )

        if not next_button.is_enabled():
            print("Next button is disabled; stopping pagination.")
            break

        driver.execute_script("arguments[0].click();", next_button)
        time.sleep(2)

    except TimeoutException:
        print("No next button found; stopping pagination.")
        break
    except Exception as e:
        print(f"Error moving to next page: {e}")
        break

driver.quit()

'''wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tr.border-base")))

soup = BeautifulSoup(driver.page_source, "html.parser")
results = soup.select("tr.border-base")

for result in results:
    a_tag = result.select_one("a")
    if not a_tag:
        continue

    relative_link = a_tag.get("href")
    full_link = BASE_DOMAIN + relative_link

    if full_link not in links:
        links.append(full_link)

print(f"Collected {len(links)} links from first page")

driver.quit()'''


############################
# SCRAPE DETAILS
############################

for detail_link in links:

    try:
        response = requests.get(detail_link, timeout=DETAIL_REQUEST_TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        title_tag = soup.select_one(
            "h2"
        )
        grant_name = title_tag.get_text(strip=True) if title_tag else None

        agency = None
        p_tag = soup.select_one("p.usa-intro")
        if p_tag:
            agency = p_tag.get_text(strip=True).replace("Agency:", "").strip()

        deadline = None
        for tag in soup.find_all("div", class_="usa-tag"):
            if "Closing:" in tag.get_text():
                span = tag.find("span")
                deadline = span.get_text(strip=True) if span else None
                break

        award_min = 0
        award_max = 0

        blocks = soup.select('div[data-testid="grid"]')

        for block in blocks:
            value_tag = block.select_one("p.font-sans-sm.text-bold")
            label_tag = block.select_one("p.desktop-lg\\:font-sans-sm")

            if not value_tag or not label_tag:
                continue

            value_text = value_tag.get_text(strip=True)
            label_text = label_tag.get_text(strip=True)

            try:
                numeric_value = int(
                    value_text.replace("$", "").replace(",", "").strip()
                )
            except ValueError:
                numeric_value = 0

            if "Minimum" in label_text:
                award_min = numeric_value
            elif "Maximum" in label_text:
                award_max = numeric_value

        description = None

        header = soup.find("h2", string=lambda x: x and "Description" in x)

        # Filter: keep only descriptions that mention at least 3 of the target phrases
        target_phrases = [
            "youth employment",
            "workforce development",
            "employability",
            "job placement",
            "job creation",
            "livelihoods",
            "economic empowerment",
            "economic inclusion",
            "apprenticeship",
            "internship",
            "mentorship",
            "job readiness",
            "job search",
            "labor market activation",
            "economic participation",
            "labor market entry",
            "neet",
            "work readiness",
            "job seekers",
            "early-career",
            "access to opportunities",
            "reducing inequalities",
            "skills development",
            "vocational training",
            "technical training",
            "soft skills",
            "digital skills",
            "vocational skills",
            "technical skills",
            "green jobs",
            "green skills",
            "tvet",
            "upskilling",
            "reskilling",
            "employability skills",
            "curriculum development",
            "vocational training center",
            "career center",
            "ai skills",
            "climate change",
            "financial literacy",
            "circular economy",
            "higher education",
            "university",
            "educational institutions",
            "life skills",
            "transversal skills",
            "entrepreneurial skills",
            "blended training",
            "entrepreneurship",
            "sme development",
            "private sector development",
            "self employment",
            "virtual jobs",
            "income generation",
            "startup incubation",
            "employer engagement",
            "business acceleration",
            "micro entrepreneurship",
            "new business creation",
            "sme",
            "incubation",
            "green entreprenurship",
            "women entrepreneurship",
            "startup",
            "startup support",
            "financial inclusion",
            "home-based businesses",
            "msme",
            "microbusiness",
            "freelance",
            "gig work",
            "gig economy",
            "capacity building",
            "systems strengthening",
            "framework",
            "action plan",
            "competitiveness",
            "skills gaps",
            "business association",
            "chamber of commerce",
            "industry federation",
        ]

        documents = [a["href"] for a in soup.select('tbody a.usa-link')]

        application_link = None

        span = soup.find("span", string="View on Grants.gov")
        if span:
            parent_a = span.find_parent("a")
            if parent_a:
                application_link = parent_a.get("href")

        if header:
            div = header.find_next("div")

            if div:
                text = div.get_text(separator="\n", strip=True)
                first_p = div.find("p")
                if first_p:
                    # Case 1: multiple <p> tags → use first one
                    description = first_p.get_text(strip=True)

                else:
                    # Case 2: no <p> tags → get full div text
                    description = div.get_text(strip=True)

                description_text = (text or "").lower()
                match_count = sum(1 for phrase in target_phrases if phrase in description_text)

                if match_count >= 1:

                    ai_summary = generate_simpler_summary(description)

                    rows.append({
                        "Date Scraped": datetime.datetime.now().strftime("%Y-%m-%d"),
                        "Grant Name": grant_name,
                        "Agency": agency,
                        "Due Date": deadline,
                        "Award Minimum": award_min,
                        "Award Maximum": award_max,
                        "Description": description,
                        "Documents": documents,
                        "Application Link": application_link,
                        "AI Summary": ai_summary
                    })

        print(f"Scraped: {grant_name}")

        time.sleep(1)

    except Exception as e:
        print("Error scraping:", e)

df = pd.DataFrame(rows)

if df.empty:
    print("No grants found; skipping Excel update and document download.")
    raise SystemExit(0)

df["Documents"] = df["Documents"].apply(json.dumps)
df = df[df["Agency"].isin(ALLOWED_AGENCIES)]


if df.empty:
    print("All grants expired; skipping Excel update.")
    raise SystemExit(0)

if os.path.exists(EXCEL_FILE):
    existing_sheets = pd.ExcelFile(EXCEL_FILE).sheet_names

    if SHEET_NAME in existing_sheets:
        existing_df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

        # Purge expired rows from the existing sheet
        before_purge = len(existing_df)
        existing_df = existing_df[existing_df["Due Date"].apply(is_not_expired)]
        purged = before_purge - len(existing_df)
        if purged:
            print(f"Removed {purged} expired grant(s) from existing sheet.")

        # Append only new (not-yet-recorded) grants
        existing_links = set(existing_df["Application Link"])
        new_rows = df[~df["Application Link"].isin(existing_links)]
        combined_df = pd.concat([existing_df, new_rows], ignore_index=True)

        # Rewrite the entire sheet so purged rows are actually gone
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined_df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

        if not new_rows.empty:
            print(f"Added {len(new_rows)} new grant(s).")
        else:
            print("No new grants.")

    else:
        # File exists but sheet doesn't — add new sheet without touching others
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        print(f"Created new sheet '{SHEET_NAME}'")

else:
    # File doesn't exist at all — create it
    df.to_excel(EXCEL_FILE, sheet_name=SHEET_NAME, index=False)
    print("Created new Excel file")


apply_impact_formatting(EXCEL_FILE, SHEET_NAME)

download_documents(
    rows=rows,
    BASE_DOWNLOAD_DIR=os.path.join(BASE_DOWNLOAD_DIR,SHEET_NAME),
    BASE_DOMAIN=BASE_DOMAIN,
    grant_name_col="Grant Name",
    docs_arr_col="Documents"
)