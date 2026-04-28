from summarizer import generate_sam_summary
import asyncio
import pandas as pd
import re
import time
from urllib.parse import urljoin
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
import os
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
from reqs import *
from datetime import datetime

BASE_URL = "https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/calls-for-proposals"

load_dotenv()
EXCEL_FILE = os.environ["EXCEL_FILE"]
def is_not_expired(deadline_str):
    if not deadline_str:
        return True
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d-%m-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(str(deadline_str).strip(), fmt) >= datetime.today()
        except ValueError:
            continue
    return True

def norm(text):
    return re.sub(r"\s+", " ", text or "").strip()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_COLOR = "1F4E79"
HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

DEFAULT_WIDTHS = {
    "Opportunity ID": 18,
    "Opportunity Type": 22,
    "Title": 40,
    "Donor Name": 28,
    "Geographic Area": 24,
    "Focus / Sector": 28,
    "Application Deadline": 20,
    "Amount Min (USD)": 18,
    "Amount Max (USD)": 18,
    "Eligibility": 42,
    "Eligibility Requirements": 45,
    "Matched Keywords": 32,
    "Source Link": 45,
    "Original Link": 45,
    "Date Posted": 18,
    "Date Scraped": 16,
    "Grant Name": 38,
    "Agency": 28,
    "Due Date": 18,
    "Award Minimum": 18,
    "Award Maximum": 18,
    "Description": 65,
    "Documents": 45,
    "Application Link": 45,
    "AI Summary": 65,
    "Post Link": 50,
    "Funding Amount (USD)": 22,
    "Grant Link": 45,
    "Deadline": 18,
}

def _auto_width(header: str) -> float:
    header = (header or "").strip()
    if not header:
        return 18
    return min(max(len(header) + 4, 18), 45)

def apply_impact_formatting(path: str, sheet_name: str, custom_widths=None):
    wb = openpyxl.load_workbook(path)

    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    widths = {**DEFAULT_WIDTHS, **(custom_widths or {})}

    if ws.max_row >= 1:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            header = "" if cell.value is None else str(cell.value)

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(
                header, _auto_width(header)
            )

        ws.row_dimensions[1].height = 30

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = BODY_ALIGNMENT

    ws.freeze_panes = "A2"
    wb.save(path)

MENA_COUNTRIES = [
    "Morocco","Algeria","Tunisia","Egypt","Jordan","Palestine","Palestinian",
    "West Bank","Gaza","Yemen","UAE","United Arab Emirates","Saudi Arabia",
    "Lebanon","Bahrain","Syria","MENA","Middle East","North Africa",
    "Arab World","GCC","Maghreb","Levant",
]

KEYWORDS = [
    "youth employment","workforce development","employability","job placement",
    "job creation","livelihoods","economic empowerment","economic inclusion",
    "apprenticeship","internship","mentorship","job readiness","job search",
    "labor market activation","economic participation","labor market entry",
    "NEET","work readiness","job seekers","early-career","reducing inequalities",
    "skills development","vocational training","technical training","soft skills",
    "digital skills","green jobs","green skills","TVET","upskilling","resciling",
    "entrepreneurship","SME development","financial inclusion","gig economy",
]

# ✅ UPDATED: Added "AI Summary" and "Full_Description" (temp storage)
COLUMNS = [
    "Opportunity ID","Title","Application Deadline","Matched Keywords",
    "Geographic Area","Original Link", "Full_Description", "AI Summary"
]

def contains_mena(text):
    if not text:
        return []
    text_lower = text.lower()
    return [c for c in MENA_COUNTRIES if c.lower() in text_lower]

def find_keywords(text):
    if not text:
        return []
    text_lower = text.lower()
    return [kw for kw in KEYWORDS if kw.lower() in text_lower]

TARGET_NEW_ENTRIES = 5

async def scrape():
    rows = []

    # Load existing links from Excel to avoid duplicates
    SHEET_NAME = "eu_comm"
    existing_links = set()
    if os.path.exists(EXCEL_FILE):
        try:
            if SHEET_NAME in pd.ExcelFile(EXCEL_FILE).sheet_names:
                existing_df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
                if "Original Link" in existing_df.columns:
                    existing_links = set(existing_df["Original Link"].dropna())
                    print(f"📋 Loaded {len(existing_links)} existing links from Excel")
        except Exception as e:
            print(f"⚠️ Could not load existing links: {e}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        )
        page = await context.new_page()

        print("🌐 Opening page...")
        await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30000)

        done = False
        for page_num in range(3):
            if done:
                break
            print(f"\n🔎 Scraping page {page_num + 1}/3\n")

            try:
                await page.wait_for_selector('div:has-text("Deadline")', timeout=10000, state="attached")
            except PlaywrightTimeout:
                print("❌ Timed out waiting for cards — stopping.")
                break

            cards = page.locator("div").filter(has_text="Deadline")
            count = await cards.count()
            print(f"   Found {count} cards on this page")

            for i in range(count):
                if done:
                    break
                try:
                    card = cards.nth(i)
                    text = await card.inner_text(timeout=5000)
                    print(f"\n---\n📌 Card {i+1}")

                    link_el = card.locator("a").first
                    if not await link_el.count():
                        print("❌ REJECTED: No link")
                        continue

                    link = await link_el.get_attribute("href", timeout=5000)
                    if not link:
                        print("❌ REJECTED: No href")
                        continue

                    full_link = urljoin(BASE_URL, link)

                    if full_link in existing_links:
                        print(f"⏭️ Already in Excel, skipping")
                        continue

                    print(f"🔗 Opening: {full_link[:80]}...")

                    # Open detail page in a new tab — listing page stays loaded
                    detail_page = await context.new_page()
                    try:
                        await detail_page.goto(full_link, wait_until="domcontentloaded", timeout=15000)
                    except PlaywrightTimeout:
                        print("⚠️ Detail page timeout, skipping")
                        await detail_page.close()
                        continue
                    except Exception as e:
                        print(f"❌ Navigation error: {e}")
                        await detail_page.close()
                        continue

                    content = await detail_page.inner_text("body", timeout=8000)

                    try:
                        search_text = await detail_page.evaluate("""() => {
                            const main = document.querySelector('main, [role="main"], article, #content');
                            if (main && main.innerText.trim().length > 200) return main.innerText;
                            const clone = document.body.cloneNode(true);
                            clone.querySelectorAll(
                                'nav, header, footer, aside, [role="navigation"], [role="complementary"], ' +
                                '[role="banner"], [role="search"], [role="menubar"], ' +
                                '[aria-label*="filter"], [aria-label*="navigation"], ' +
                                '[aria-label*="sidebar"], [aria-label*="menu"]'
                            ).forEach(el => el.remove());
                            return clone.innerText;
                        }""")
                        if not search_text or len(search_text.strip()) < 100:
                            search_text = content
                    except Exception:
                        search_text = content

                    await detail_page.close()

                    mena_matches = contains_mena(search_text)
                    if not mena_matches:
                        print("❌ REJECTED: No MENA match")
                        continue
                    print(f"✅ MENA: {mena_matches}")

                    keyword_matches = find_keywords(search_text)
                    if not keyword_matches:
                        print("❌ REJECTED: No keyword match")
                        continue
                    print(f"✅ Keywords: {keyword_matches}")

                    deadline = ""
                    deadline_match = re.search(r'Deadline[:\s]+([^\n]+)', text, re.I)
                    if deadline_match:
                        deadline = deadline_match.group(1).strip()

                    rows.append({
                        "Opportunity ID": "",
                        "Title": text.split("\n")[0].strip(),
                        "Application Deadline": deadline,
                        "Matched Keywords": ", ".join(keyword_matches),
                        "Geographic Area": ", ".join(mena_matches),
                        "Original Link": full_link,
                        "Full_Description": content,
                        "AI Summary": ""
                    })
                    existing_links.add(full_link)
                    print(f"✅ New entry {len(rows)}/{TARGET_NEW_ENTRIES} collected")

                    if len(rows) >= TARGET_NEW_ENTRIES:
                        print(f"\n🎯 Reached {TARGET_NEW_ENTRIES} new entries — stopping.")
                        done = True
                        break

                except Exception as e:
                    print(f"❌ Error processing card {i+1}: {e}")
                    continue

            if done:
                break

            # Pagination
            print("🔄 Checking for next page...")
            next_button = page.locator("button:has-text('Next'):not(:disabled)")
            if await next_button.count() > 0 and await next_button.is_visible(timeout=3000):
                print("➡️ Moving to next page...")
                try:
                    await next_button.click(timeout=5000)
                    await page.wait_for_selector('div:has-text("Deadline")', timeout=10000, state="attached")
                except Exception as e:
                    print(f"⚠️ Could not go to next page: {e}")
                    break
            else:
                print("⛔ No more pages")
                break

        await browser.close()

    if not rows:
        print("\n⚠️ No new grants found in the first 3 pages.")
        return pd.DataFrame(columns=COLUMNS)
    
    if rows:
        print(f"\n🤖 Starting AI Summarization for {len(rows)} opportunities...")
        
        for idx, row in enumerate(rows):
            print(f"   Processing {idx+1}/{len(rows)}: {row['Title'][:40]}...")
            
            # Prepare data structure expected by generate_sam_summary
            # We map your EU columns to the keys the function expects
            opp_data = {
                "Title": row.get("Title", ""),
                "Donor Name": "European Commission",
                "Geographic Area": row.get("Geographic Area", ""),
                "Focus / Sector": "", 
                "Eligibility": "",
                "Amount Max (USD)": "",
                "Application Deadline": row.get("Application Deadline", ""),
                "body": row.get("Full_Description", "") # Pass the full scraped text
            }

            try:
                summary = generate_sam_summary(opp_data)
                row["AI Summary"] = summary
            except Exception as e:
                print(f"      ❌ Error generating summary: {e}")
                row["AI Summary"] = "Error generating summary."
            
            #  RATE LIMIT PROTECTION: Pause between calls
            time.sleep(1.5) 

        print("✅ All summaries generated.\n")

    SHEET_NAME = "eu_comm"
    if rows:
        df = pd.DataFrame(rows, columns=COLUMNS)
        print(f"\n✅ Done. Saved {len(rows)} opportunities")
        
        # Drop the temporary Full_Description column before saving to Excel 
        # (Optional: Remove this line if you want the full text in Excel too)
        if "Full_Description" in df.columns:
            df = df.drop(columns=["Full_Description"])
            # Ensure COLUMNS matches for the final save logic if needed, 
            # but passing the DF directly is safer.
            final_columns = [c for c in COLUMNS if c != "Full_Description"]
            df = df[final_columns]
        
        df = df[df["Application Deadline"].apply(is_not_expired)]
        print("Filtered expired deadlines.")

        if os.path.exists(EXCEL_FILE):
            if SHEET_NAME in pd.ExcelFile(EXCEL_FILE).sheet_names:
                existing_df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

                before_purge = len(existing_df)
                existing_df = existing_df[existing_df["Application Deadline"].apply(is_not_expired)]
                purged = before_purge - len(existing_df)
                if purged:
                    print(f"Removed {purged} expired grant(s) from existing sheet.")

                combined = pd.concat([existing_df, df], ignore_index=True)
                combined = combined.drop_duplicates(subset=["Original Link"], keep="last")

                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    combined.to_excel(writer, sheet_name=SHEET_NAME, index=False)
                print(f"Merged and saved. Total rows: {len(combined)}")
            else:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer: 
                    df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
                print(f"Created new sheet '{SHEET_NAME}'")
        else:
            df.to_excel(EXCEL_FILE, sheet_name=SHEET_NAME, index=False)
            print("Created new Excel file")

        apply_impact_formatting(EXCEL_FILE, SHEET_NAME)
        
        print(f"\n{'='*50}\n  Done — {len(df)} matching rows\n{'='*50}\n")
        pd.set_option("display.max_columns", None)
        pd.set_option("display.max_colwidth", 60)
        pd.set_option("display.width", 220)
        print(df.to_string(index=False))        
        
        return df
    else:
        print("\n⚠️ No data collected. Check filters or website structure.")
        return pd.DataFrame(columns=COLUMNS)

if __name__ == "__main__":
    asyncio.run(scrape())