'''import subprocess
import traceback
from datetime import datetime
from pathlib import Path
import os

from upload_to_sharepoint import process_uploads

scrapers = ["attempt2.py", "scraper.py", "impact_funding_scraper.py", "dev_aid.py", "eu_comm.py", "fundsforngos_webscraper.py", "sam_fast.py"]  # change to test later

_ROOT = Path(__file__).resolve().parent
_LOG_DIR = _ROOT / "logs"
_LOG_DIR.mkdir(exist_ok=True)
_error_log_path = _LOG_DIR / f"scraper_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"


def _log_error(section: str) -> None:
    with open(_error_log_path, "a", encoding="utf-8") as log:
        log.write(section)
        if not section.endswith("\n"):
            log.write("\n")
        log.write("\n")


with open(_error_log_path, "w", encoding="utf-8") as log:
    log.write(f"Run started: {datetime.now().isoformat()}\n")
    log.write("=" * 60 + "\n\n")

try:
    for scraper in scrapers:
        try:
            result = subprocess.run(
                ["python", scraper],
                cwd=_ROOT,
                capture_output=True,
                text=True,
            )
            if result.returncode == 0:
                print(f"{scraper} success")
            else:
                _log_error(
                    f"[{scraper}] FAILED exit code {result.returncode}\n"
                    f"--- STDOUT ---\n{result.stdout}\n"
                    f"--- STDERR ---\n{result.stderr}\n"
                )
                print(f"{scraper} failed — details written to {_error_log_path}")
        except Exception as e:
            _log_error(f"[{scraper}] CRASHED: {e!r}\n{traceback.format_exc()}")
            print(f"{scraper} crashed — details written to {_error_log_path}")
    # call upload to sharepoint here
    process_uploads()
    print("Uploaded to Sharepoint Successfully")
except Exception as e:
    _log_error(f"SharePoint upload error: {e!r}\n{traceback.format_exc()}")
    print(f"Error uploading to sharepoint — details written to {_error_log_path}")'''

import os
import sys
import subprocess
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from upload_to_sharepoint import process_uploads

scrapers = [
    "attempt2.py",
    "scraper.py",
    "impact_funding_scraper.py",
    "dev_aid.py",
    "eu_comm.py",
    "fundsforngos_webscraper.py",
    "sam_fast.py",
]

SCRAPER_COMMANDS = {
    "attempt2.py": [sys.executable, "attempt2.py"],
    "scraper.py": [sys.executable, "scraper.py"],
    "impact_funding_scraper.py": [sys.executable, "impact_funding_scraper.py"],
    "dev_aid.py": [sys.executable, "dev_aid.py", "--headless", "--pages", "3"],
    "eu_comm.py": [sys.executable, "eu_comm.py"],
    "fundsforngos_webscraper.py": [sys.executable, "fundsforngos_webscraper.py"],
    "sam_fast.py": [sys.executable, "sam_fast.py", "--headless", "--pages", "2"],
}

SCRAPER_TIMEOUTS = {
    "attempt2.py": 7200,
    "scraper.py": 7200,
    "impact_funding_scraper.py": 7200,
    "dev_aid.py": 7200,
    "eu_comm.py": 7200,
    "fundsforngos_webscraper.py": 7200,
    "sam_fast.py": 7200,
}

_ROOT = Path(__file__).resolve().parent
_LOG_DIR = _ROOT / "logs"
_LOG_DIR.mkdir(exist_ok=True)

_error_log_path = _LOG_DIR / f"scraper_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"


def _log_error(section: str) -> None:
    with open(_error_log_path, "a", encoding="utf-8") as log:
        log.write(section)
        if not section.endswith("\n"):
            log.write("\n")
        log.write("\n")


def _get_excel_path() -> Path | None:
    excel_name = os.environ.get("EXCEL_FILE")
    if not excel_name:
        return None

    excel_path = Path(excel_name)
    if not excel_path.is_absolute():
        excel_path = _ROOT / excel_path

    return excel_path


def _print_workbook_state(scraper_name: str) -> None:
    excel_path = _get_excel_path()

    if excel_path is None:
        print(f"[{scraper_name}] EXCEL_FILE is not set")
        return

    if not excel_path.exists():
        print(f"[{scraper_name}] Excel file does not exist yet: {excel_path}")
        return

    try:
        wb = load_workbook(excel_path)
        print(f"[{scraper_name}] Excel file: {excel_path}")
        print(f"[{scraper_name}] Sheets now: {wb.sheetnames}")
    except Exception as e:
        print(f"[{scraper_name}] Could not inspect workbook: {e!r}")


with open(_error_log_path, "w", encoding="utf-8") as log:
    log.write(f"Run started: {datetime.now().isoformat()}\n")
    log.write("=" * 60 + "\n\n")

try:
    for scraper in scrapers:
        print(f"\n{'=' * 70}")
        print(f"RUNNING {scraper}")
        print(f"{'=' * 70}")

        try:
            result = subprocess.run(
                SCRAPER_COMMANDS[scraper],
                cwd=_ROOT,
                text=True,
                timeout=SCRAPER_TIMEOUTS[scraper],
            )

            if result.stdout:
                print(f"\n--- {scraper} STDOUT ---")
                print(result.stdout)

            if result.stderr:
                print(f"\n--- {scraper} STDERR ---")
                print(result.stderr)

            if result.returncode == 0:
                print(f"{scraper} success")
            else:
                _log_error(
                    f"[{scraper}] FAILED exit code {result.returncode}\n"
                    f"--- STDOUT ---\n{result.stdout}\n"
                    f"--- STDERR ---\n{result.stderr}\n"
                )
                print(f"{scraper} failed — details written to {_error_log_path}")

            _print_workbook_state(scraper)


        except subprocess.TimeoutExpired as e:
            _log_error(f"[{scraper}] TIMED OUT\n--- STDOUT ---\n{e.stdout}\n--- STDERR ---\n{e.stderr}\n")
            print(f"{scraper} timed out — details written to {_error_log_path}")

    print("\n" + "=" * 70)
    print("STARTING SHAREPOINT UPLOAD")
    print("=" * 70)

    process_uploads()
    print("Uploaded to Sharepoint Successfully")

except Exception as e:
    _log_error(f"SharePoint upload error: {e!r}\n{traceback.format_exc()}")
    print(f"Error uploading to sharepoint — details written to {_error_log_path}")
